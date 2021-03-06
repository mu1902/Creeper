''' 爬取巨潮财务报告 '''
import datetime
import json
import os
import random
import re
import time
import sys

import openpyxl

import Creeper as cp
from pyquery import PyQuery as pq

DIR = os.path.dirname(os.path.abspath(__file__))
url = 'http://www.cninfo.com.cn/new/hisAnnouncement/query'
dq_type = ['category_yjdbg_szsh', 'category_bndbg_szsh',
           'category_sjdbg_szsh', 'category_ndbg_szsh']
yg_type = ['一', '半', '三']
today = datetime.date.today().strftime('%Y-%m-%d')
day = (datetime.date.today() + datetime.timedelta(days=1)).strftime('%Y-%m-%d')
# day = '2020-08-11'

dqbg_codes = []
dqbg_time = []
param = {'pageNum': 1,
         'pageSize': 30,
         'column': 'szse',
         'tabName': 'fulltext',
         'plate': '',
         'stock': '',
         'searchkey': '',
         'secid': '',
         'category': dq_type[1],
         'trade': '',
         'seDate': today+'~'+day,
         'sortName': '',
         'sortType': '',
         'isHLtitle': True}

dqbg = cp.downloader.get_html(url, param, method='post')
pages = json.loads(dqbg)['totalpages'] + 1
codes = json.loads(dqbg)['announcements']
for c in codes if codes else []:
    if c['secCode'] not in dqbg_codes:
        dqbg_codes.append(c['secCode'])
        dqbg_time.append(datetime.datetime.fromtimestamp(
            c['announcementTime']/1000).strftime("%Y-%m-%d %H:%M:%S"))
for i in range(2, pages+1):
    param['pageNum'] = i
    dqbg = cp.downloader.get_html(url, param, method='post')
    codes = json.loads(dqbg)['announcements']
    for c in codes if codes else []:
        if c['secCode'] not in dqbg_codes:
            dqbg_codes.append(c['secCode'])
            dqbg_time.append(datetime.datetime.fromtimestamp(
                c['announcementTime']/1000).strftime("%Y-%m-%d %H:%M:%S"))

yg_codes = []
kb_codes = []
yg_time = []
kb_time = []
param = {'pageNum': 1,
         'pageSize': 30,
         'column': 'szse',
         'tabName': 'fulltext',
         'plate': '',
         'stock': '',
         'searchkey': '',
         'secid': '',
         'category': 'category_yjygjxz_szsh',
         'trade': '',
         'seDate': today+'~'+day,
         'sortName': '',
         'sortType': '',
         'isHLtitle': True}

yg = cp.downloader.get_html(url, param, method='post')
pages = json.loads(yg)['totalpages'] + 1
codes = json.loads(yg)['announcements']
for c in codes if codes else []:
    if '预' in c['announcementTitle']:
        yg_codes.append(c['secCode'])
        yg_time.append(datetime.datetime.fromtimestamp(
            c['announcementTime']/1000).strftime("%Y-%m-%d %H:%M:%S"))
    else:
        kb_codes.append(c['secCode'])
        kb_time.append(datetime.datetime.fromtimestamp(
            c['announcementTime']/1000).strftime("%Y-%m-%d %H:%M:%S"))
for i in range(2, pages+1):
    param['pageNum'] = i
    yg = cp.downloader.get_html(url, param, method='post')
    codes = json.loads(yg)['announcements']
    for c in codes if codes else []:
        if '预' in c['announcementTitle']:
            yg_codes.append(c['secCode'])
            yg_time.append(datetime.datetime.fromtimestamp(
                c['announcementTime']/1000).strftime("%Y-%m-%d %H:%M:%S"))
        else:
            kb_codes.append(c['secCode'])
            kb_time.append(datetime.datetime.fromtimestamp(
                c['announcementTime']/1000).strftime("%Y-%m-%d %H:%M:%S"))

print(len(dqbg_codes), len(yg_codes), len(kb_codes))


workbook = openpyxl.load_workbook(DIR+'\\业绩报告.xlsx')
sheet1 = workbook['2020年三季度预告']
sheet2 = workbook['2020年半年快报']
sheet3 = workbook['2020年半年报']


for i in range(len(yg_codes)):
    code = yg_codes[i]+'.SH' if yg_codes[i][0] == '6' else yg_codes[i]+'.SZ'
    sheet1.cell(row=i+4, column=1, value=code)
    sheet1.cell(row=i+4, column=15, value=yg_time[i])
    # for j in range(2, 14):
    #     sheet1.cell(row=i+4, column=j, value=sheet1.cell(row=4, column=j).value)

for i in range(len(kb_codes)):
    code = kb_codes[i]+'.SH' if kb_codes[i][0] == '6' else kb_codes[i]+'.SZ'
    sheet2.cell(row=i+4, column=1, value=code)
    sheet2.cell(row=i+4, column=19, value=kb_time[i])

for i in range(len(dqbg_codes)):
    code = dqbg_codes[i] + \
        '.SH' if dqbg_codes[i][0] == '6' else dqbg_codes[i]+'.SZ'
    sheet3.cell(row=i+4, column=1, value=code)
    sheet3.cell(row=i+4, column=19, value=dqbg_time[i])

workbook.save(DIR+'\\业绩报告'+day+'.xlsx')
cp.tool.send_email([], '业绩报告', '', file=DIR+'\\业绩报告'+day+'.xlsx')
os.remove(DIR+'\\业绩报告'+day+'.xlsx')
